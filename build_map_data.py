#!/usr/bin/env python3
"""Bangun map_province_data.json: perusahaan per provinsi + sektor BKPM & PDB + laba terbaru + PDRB 2026 per Sektor."""
from __future__ import annotations

import json
import math
import re
from pathlib import Path
import openpyxl

import pandas as pd

ROOT = Path(__file__).resolve().parent
DETAIL_CSV = Path(__file__).resolve().parents[1] / "scrap_idx" / "detail_perusahaan.csv"
ALL_LK = ROOT / "all_lk.csv"
PRIMARY_MAP = ROOT / "bkpm_subindustri_primary.csv"
GEOJSON = ROOT / "dashboard/public/38 Provinsi Indonesia - Provinsi.json"
OUT_JSON = ROOT / "dashboard/public/map_province_data.json"

BKPM_SECTORS = [
    "ENERGI", "KEUANGAN", "KONSTRUKSI", "PARIWISATA", "PENGANGKUTAN",
    "PERDAGANGAN", "PERIKANAN", "PERINDUSTRIAN", "PERTAMBANGAN", "PERTANIAN",
]

ALIASES = {
    "daerah khusus ibukota jakarta": "DKI Jakarta",
    "dki jakarta": "DKI Jakarta",
    "jakarta": "DKI Jakarta",
    "daerah istimewa yogyakarta": "Daerah Istimewa Yogyakarta",
    "di yogyakarta": "Daerah Istimewa Yogyakarta",
    "yogyakarta": "Daerah Istimewa Yogyakarta",
    "kepulauan bangka belitung": "Kepulauan Bangka Belitung",
    "bangka belitung": "Kepulauan Bangka Belitung",
}


def parse_value(val):
    if pd.isna(val) or val == "-" or val == "":
        return None
    s = str(val).strip().replace(",", "")
    negative = s.startswith("(") and s.endswith(")")
    if negative:
        s = s[1:-1]
    multiplier = 1
    if s.endswith("T"):
        multiplier = 1e12
        s = s[:-1]
    elif s.endswith("B"):
        multiplier = 1e9
        s = s[:-1]
    elif s.endswith("M"):
        multiplier = 1e6
        s = s[:-1]
    try:
        num = float(s) * multiplier
        return -num if negative else num
    except ValueError:
        return None


def clean_province(prov) -> str | None:
    if prov is None:
        return None
    try:
        if pd.isna(prov):
            return None
    except (TypeError, ValueError):
        pass
    s = str(prov).strip()
    return s if s and s.lower() != "nan" else None


def sanitize_for_json(obj):
    """Ganti NaN/Inf agar output JSON valid di browser."""
    if isinstance(obj, dict):
        return {k: sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [sanitize_for_json(v) for v in obj]
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    try:
        if pd.isna(obj):
            return None
    except (TypeError, ValueError):
        pass
    return obj


def load_geo_provinces() -> list[str]:
    with open(GEOJSON, encoding="utf-8") as f:
        geo = json.load(f)
    return [feat["properties"]["PROVINSI"] for feat in geo["features"]]


def normalize(name: str | None, geo_provinces: list[str]) -> str | None:
    if not name:
        return None
    n = re.sub(r"\s+", " ", str(name).strip(" -,."))
    low = n.lower()
    if low in ALIASES:
        return ALIASES[low]
    for g in geo_provinces:
        if n.lower() == g.lower():
            return g
    return n


def extract_province(alamat: str | None, geo_provinces: list[str]) -> str | None:
    if pd.isna(alamat):
        return None
    s = str(alamat)
    m = re.search(r"Provinsi\s+([^,\d]+?)(?:\s*\d|,|$)", s, re.I)
    if m:
        return normalize(m.group(1), geo_provinces)
    low = s.lower()
    for g in sorted(geo_provinces, key=len, reverse=True):
        if g.lower() in low:
            return g
    for alias, canonical in ALIASES.items():
        if alias in low:
            return canonical
    if re.search(r"\bjakarta\b", low) and "jawa barat" not in low:
        return "DKI Jakarta"
    if re.search(r"\bbanten\b", low) or "tangerang" in low:
        return "Banten"
    if re.search(r"\bbekasi\b", low) or "cikarang" in low or "depok" in low or "bogor" in low:
        return "Jawa Barat"
    if "surabaya" in low or "sidoarjo" in low:
        return "Jawa Timur"
    if "bandung" in low:
        return "Jawa Barat"
    if "semarang" in low:
        return "Jawa Tengah"
    if "medan" in low:
        return "Sumatera Utara"
    if "makassar" in low:
        return "Sulawesi Selatan"
    if "denpasar" in low:
        return "Bali"
    return None


def normalize_province_bkpm(name: str | None, geo_provinces: list[str]) -> str | None:
    if not name:
        return None
    s = str(name).strip().lower()
    s = s.replace("-", " ")
    s = s.replace("kep.", "kepulauan")
    for g in geo_provinces:
        g_clean = g.lower().replace("-", " ")
        if s == g_clean or s in g_clean or g_clean in s:
            return g
    return name


def normalize_sector_name(name):
    if not name:
        return ""
    s = str(name).strip().lower()
    s = s.replace(";", ",").replace(" dan ", " & ")
    s = re.sub(r"[,\s]+", " ", s)
    return s


def main():
    geo_provinces = load_geo_provinces()
    
    # 1. Load PDRB from PDRB_2026.xlsx
    pdrb_path = "/Users/67620/scrap_sector/dashboard-lapkeu/PDRB_2026.xlsx"
    pdrb_stats = {}
    national_sector_totals = {}
    if Path(pdrb_path).exists():
        print(f"Reading PDRB 2026 Excel: {pdrb_path}")
        wb_pdrb = openpyxl.load_workbook(pdrb_path, read_only=True)
        sheet_pdrb = wb_pdrb.active
        rows_pdrb = list(sheet_pdrb.iter_rows(values_only=True))
        if len(rows_pdrb) > 0:
            headers = [str(h).strip() for h in rows_pdrb[0]]
            headers_lower = [h.lower() for h in headers]
            
            # Sektor columns (between Provinsi and Total PDRB)
            # Last column (headers[-1]) is Total PDRB
            sector_headers = headers[1:-1]
            total_header_idx = len(headers) - 1
            
            for r in rows_pdrb[1:]:
                if not r or r[0] is None:
                    continue
                prov_name_raw = str(r[0]).strip()
                prov_name = normalize_province_bkpm(prov_name_raw, geo_provinces)
                
                r_padded = list(r) + [None] * (len(headers) - len(r))
                
                # Total PDRB
                total_val_raw = r_padded[total_header_idx]
                if total_val_raw is None:
                    continue
                
                try:
                    total_pdrb = int(float(str(total_val_raw).replace(",", ".")) * 1000000)
                except ValueError:
                    continue
                
                # Sectors PDRB values
                sectors_data = []
                for s_idx, sec_name in enumerate(sector_headers):
                    col_val = r_padded[s_idx + 1]
                    if col_val is not None:
                        try:
                            sec_val = int(float(str(col_val).replace(",", ".")) * 1000000)
                            national_sector_totals[sec_name] = national_sector_totals.get(sec_name, 0) + sec_val
                            share = sec_val / total_pdrb if total_pdrb > 0 else 0
                            sectors_data.append({
                                "sector": sec_name,
                                "value": sec_val,
                                "share": share
                            })
                        except ValueError:
                            pass
                
                # Sort sectors in descending order and select top 5
                sectors_data.sort(key=lambda x: -x["value"])
                top_5 = sectors_data[:5]
                
                pdrb_stats[prov_name] = {
                    "pdrb": total_pdrb,
                    "pdrbYear": "2026",
                    "top5Sectors": top_5,
                    "sectors": {s["sector"]: s["value"] for s in sectors_data}
                }
        print(f"Successfully loaded PDRB 2026 data for {len(pdrb_stats)} provinces.")
    else:
        print(f"Warning: PDRB 2026 Excel not found at {pdrb_path}.")

    # 2. Load BKPM Excel for UMR and other regional indicators
    bkpm_path = "/Users/67620/scrap_sector/dashboard-lapkeu/bkpm-data.xlsx"
    bkpm_stats = {}
    if Path(bkpm_path).exists():
        print(f"Reading BKPM Excel: {bkpm_path}")
        wb = openpyxl.load_workbook(bkpm_path, read_only=True)
        sheet = wb['Sheet1']
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) > 0:
            headers = rows[0]
            headers_lower = [str(h).lower() for h in headers]
            
            for r in rows[1:]:
                if not r or r[0] is None:
                    continue
                prov_name_raw = str(r[0]).strip()
                prov_name = normalize_province_bkpm(prov_name_raw, geo_provinces)
                
                r_padded = list(r) + [None] * (len(headers) - len(r))
                
                # Consolidate UMR
                umr_val = None
                umr_yr = None
                for yr in ['2026', '2025']:
                    col_name = f'umr {yr}'
                    if col_name in headers_lower:
                        col_idx = headers_lower.index(col_name)
                        col_val = r_padded[col_idx]
                        if col_val is not None:
                            umr_val = str(col_val).strip()
                            umr_yr = yr
                            break
                            
                # Kawasan Industri
                kawasan = None
                if 'kawasan industri' in headers_lower:
                    kawasan = r_padded[headers_lower.index('kawasan industri')]
                    if kawasan is not None:
                        kawasan = str(kawasan).strip()
                        
                # Peluang
                peluang = None
                if 'peluang' in headers_lower:
                    peluang = r_padded[headers_lower.index('peluang')]
                    if peluang is not None:
                        peluang = str(peluang).strip()
                        
                # Penduduk
                penduduk = None
                if 'jumlah penduduk' in headers_lower:
                    penduduk = r_padded[headers_lower.index('jumlah penduduk')]
                    if penduduk is not None:
                        penduduk = str(penduduk).strip()
                        
                # Sectors
                sectors_txt = None
                if 'sectors' in headers_lower:
                    sectors_txt = r_padded[headers_lower.index('sectors')]
                    if sectors_txt is not None:
                        sectors_txt = str(sectors_txt).strip()
                        
                pstats = pdrb_stats.get(prov_name, {})
                bkpm_stats[prov_name] = {
                    "pdrb": pstats.get("pdrb"),
                    "pdrbYear": pstats.get("pdrbYear", "2026"),
                    "top5Sectors": pstats.get("top5Sectors", []),
                    "umr": umr_val,
                    "umrYear": umr_yr,
                    "kawasanIndustri": kawasan,
                    "peluang": peluang,
                    "jumlahPenduduk": penduduk,
                    "bkpmSectorsText": sectors_txt
                }
        print(f"Successfully loaded economic stats for {len(bkpm_stats)} provinces.")
    else:
        print(f"Warning: BKPM Excel not found at {bkpm_path}.")
        # Populate bkpm_stats using pdrb_stats only if bkpm is missing
        for prov_name, pstats in pdrb_stats.items():
            bkpm_stats[prov_name] = {
                "pdrb": pstats["pdrb"],
                "pdrbYear": pstats["pdrbYear"],
                "top5Sectors": pstats["top5Sectors"],
                "umr": None,
                "umrYear": None,
                "kawasanIndustri": None,
                "peluang": None,
                "jumlahPenduduk": None,
                "bkpmSectorsText": None
            }

    detail = pd.read_csv(DETAIL_CSV)
    detail["province"] = detail["Alamat Kantor"].apply(
        lambda a: extract_province(a, geo_provinces)
    )

    # 3. Load mapping_subindustri_sektor_subsektor.xlsx
    mapping_path = "/Users/67620/scrap_sector/dashboard-lapkeu/mapping_subindustri_sektor_subsektor.xlsx"
    if Path(mapping_path).exists():
        print(f"Reading mapping Excel: {mapping_path}")
        mapping_df = pd.read_excel(mapping_path)
        mapping_df = mapping_df.rename(columns={
            "Subindustri_idx": "Subindustri",
            "bkpm_sector": "bkpm_sector",
            "sektor_pdb": "sektor_pdb",
            "subsektor_pdb": "subsektor_pdb"
        })
        detail = detail.merge(mapping_df[["Subindustri", "bkpm_sector", "sektor_pdb", "subsektor_pdb"]], on="Subindustri", how="left")
    else:
        print(f"Warning: Mapping Excel not found at {mapping_path}. Using fallback CSV.")
        primary = pd.read_csv(PRIMARY_MAP)
        detail = detail.merge(primary[["Subindustri", "bkpm_sector"]], on="Subindustri", how="left")
        detail["sektor_pdb"] = None
        detail["subsektor_pdb"] = None

    lk = pd.read_csv(ALL_LK, dtype=str)
    year_cols = [c for c in lk.columns if c.isdigit()]
    latest_year = max(int(c) for c in year_cols if int(c) <= 2025)

    annual = lk[lk["Period"] == "Annualised"].copy()
    annual["NetIncome"] = annual[str(latest_year)].apply(parse_value)
    profit = annual[["Ticker", "NetIncome"]].dropna()

    detail = detail.rename(columns={"Kode": "Ticker", "Nama": "NamaPerusahaan"})
    df = detail.merge(profit, on="Ticker", how="left")

    companies_by_province: dict[str, list] = {}
    province_stats: dict[str, dict] = {}

    for prov in geo_provinces:
        pstats = bkpm_stats.get(prov, {})
        province_stats[prov] = {
            "totalCompanies": 0, 
            "byBkpmSector": {s: 0 for s in BKPM_SECTORS},
            "pdrb": pstats.get("pdrb"),
            "pdrbYear": pstats.get("pdrbYear"),
            "top5Sectors": pstats.get("top5Sectors", []),
            "umr": pstats.get("umr"),
            "umrYear": pstats.get("umrYear"),
            "kawasanIndustri": pstats.get("kawasanIndustri"),
            "peluang": pstats.get("peluang"),
            "jumlahPenduduk": pstats.get("jumlahPenduduk"),
            "bkpmSectorsText": pstats.get("bkpmSectorsText")
        }
        companies_by_province[prov] = []

    unlocated = []
    for _, row in df.iterrows():
        prov = clean_province(row["province"])
        net_income = None
        if pd.notna(row["NetIncome"]):
            ni = float(row["NetIncome"])
            if not (math.isnan(ni) or math.isinf(ni)):
                net_income = ni
        
        company = {
            "Ticker": str(row["Ticker"]),
            "NamaPerusahaan": str(row["NamaPerusahaan"]) if pd.notna(row["NamaPerusahaan"]) else "",
            "Sektor": str(row["Sektor"]) if pd.notna(row["Sektor"]) else "",
            "Subsektor": str(row["Subsektor"]) if pd.notna(row["Subsektor"]) else "",
            "Industri": str(row["Industri"]) if pd.notna(row["Industri"]) else "",
            "Subindustri": str(row["Subindustri"]) if pd.notna(row["Subindustri"]) else "",
            "bkpm_sector": str(row["bkpm_sector"]) if pd.notna(row["bkpm_sector"]) else None,
            "sektor_pdb": str(row["sektor_pdb"]) if pd.notna(row["sektor_pdb"]) else None,
            "subsektor_pdb": str(row["subsektor_pdb"]) if pd.notna(row["subsektor_pdb"]) else None,
            "NetIncome": net_income,
            "province": prov,
        }
        if not prov or prov not in companies_by_province:
            unlocated.append(company)
            continue
        companies_by_province[prov].append(company)
        province_stats[prov]["totalCompanies"] += 1
        bkpm = company["bkpm_sector"]
        if bkpm in province_stats[prov]["byBkpmSector"]:
            province_stats[prov]["byBkpmSector"][bkpm] += 1

    for prov in geo_provinces:
        companies_by_province[prov].sort(
            key=lambda c: (c["NetIncome"] is None, -(c["NetIncome"] or 0))
        )

    def get_national_total(sec):
        norm = normalize_sector_name(sec)
        for s_name, s_val in national_sector_totals.items():
            if normalize_sector_name(s_name) == norm:
                return s_val
        return 0

    pdb_sectors = sorted(
        list(set(
            str(row["sektor_pdb"]).strip() 
            for _, row in df.iterrows() 
            if pd.notna(row["sektor_pdb"]) and str(row["sektor_pdb"]).strip() != ""
        )),
        key=lambda sec: -get_national_total(sec)
    )

    payload = {
        "latestYear": latest_year,
        "bkpmSectors": BKPM_SECTORS,
        "pdbSectors": pdb_sectors,
        "provinceStats": province_stats,
        "companiesByProvince": companies_by_province,
        "unlocatedCompanies": unlocated,
        "meta": {
            "totalCompanies": int(len(df)),
            "locatedCompanies": int(df["province"].notna().sum()),
            "unlocatedCount": len(unlocated),
        },
    }

    payload = sanitize_for_json(payload)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, allow_nan=False)
    print(f"Saved {OUT_JSON}")
    print(f"  Located: {payload['meta']['locatedCompanies']} / {payload['meta']['totalCompanies']}")
    top = sorted(
        ((p, province_stats[p]["totalCompanies"]) for p in geo_provinces),
        key=lambda x: -x[1],
    )[:5]
    print("  Top provinces:", top)


if __name__ == "__main__":
    main()
