import openpyxl
from pathlib import Path

xlsx_path = Path("/Users/67620/scrap_sector/dashboard-lapkeu/pdrb-data.xlsx")
if xlsx_path.exists():
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames[:3]:
        sheet = wb[name]
        print(f"\nSheet '{name}' - first 10 rows:")
        for r_idx, r in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx < 10:
                print(r)
            else:
                break
else:
    print("pdrb-data.xlsx not found at path.")
