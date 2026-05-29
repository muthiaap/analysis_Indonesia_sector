import openpyxl
from pathlib import Path

xlsx_path = Path("/Users/67620/scrap_sector/dashboard-lapkeu/PDRB_2026.xlsx")
if xlsx_path.exists():
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\nSheet '{name}' - first 10 rows:")
        for r_idx, r in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx < 10:
                print(r)
            else:
                break
else:
    print("PDRB_2026.xlsx not found.")
