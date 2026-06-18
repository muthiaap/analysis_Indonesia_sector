import openpyxl
from pathlib import Path

ROOT = Path("/Users/67620/scrap_sector/dashboard-lapkeu")
PDB_NAT = ROOT / "PDB data 2025 2026.xlsx"
PDRB = ROOT / "PDRB_2026.xlsx"

print("PDB National Excel (PDB data 2025 2026.xlsx):")
wb = openpyxl.load_workbook(PDB_NAT, read_only=True)
print("Sheets:", wb.sheetnames)
if 'Sheet3' in wb.sheetnames:
    sheet = wb['Sheet3']
    print("Sectors in Sheet3:")
    for r_idx, r in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        if r and r[0] is not None:
            print(f"  - '{r[0]}'")

print("\nPDRB Excel (PDRB_2026.xlsx):")
wb_pdrb = openpyxl.load_workbook(PDRB, read_only=True)
headers = [str(h).strip() for h in next(wb_pdrb.active.iter_rows(values_only=True))]
print("Columns in active sheet:")
for h in headers[1:-1]:
    print(f"  - '{h}'")
